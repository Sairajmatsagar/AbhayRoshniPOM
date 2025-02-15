import openpyxl


def get_row_count(path, sheet_name):
    workbook = openpyxl.load_workbook(path)
    sheet = workbook[sheet_name]
    return sheet.max_row


def get_column_count(path, sheet_name):
    workbook = openpyxl.load_workbook(path)
    sheet = workbook[sheet_name]
    return sheet.max_column


def get_cell_data(path, sheet_name, row_number, column_number):
    workbook = openpyxl.load_workbook(path)
    sheet = workbook[sheet_name]
    return sheet.cell(row=row_number, column=column_number).value


def set_cell_data(path, sheet_name, row_number, column_number, data):
    workbook = openpyxl.load_workbook(path)
    sheet = workbook[sheet_name]
    sheet.cell(row=row_number, column=column_number).value = data
    workbook.save(path)


def get_data_from_excel(path, sheet_name):
    workbook = openpyxl.load_workbook(path)
    sheet = workbook[sheet_name]
    total_rows = sheet.max_row
    total_column = sheet.max_column
    final_list = []

    for r in range(2, total_rows+1):
        row_list = []
        for c in range(1, total_column+1):
            row_list.append(sheet.cell(row=r, column=c).value)
        final_list.append(row_list)
    return final_list


def read_data_from_exel_file(path, sheet_name):
    workbook = openpyxl.load_workbook(path)
    testdata_sheet = workbook[sheet_name]
    max_rows = testdata_sheet.max_row
    max_column = testdata_sheet.max_column
    final_list = []
    for row in range(2, max_rows+1):
        sub_list = []
        for col in range(1, max_column+1):
            sub_list.append(testdata_sheet.cell(row, col).value)
        final_list.append(sub_list)
    return final_list





